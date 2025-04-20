#############################################
#TDC 3-25-25
#
##############################################
from tkinter import *
import tkinter.font
from plcconn import *
import re
from pycomm3 import LogixDriver, CIPDriver
from datetime import datetime
import os
from openpyxl import Workbook
import openpyxl
from tkinter import ttk



win = Tk()
win.configure(bg='dark grey')
win.title("datacollection")
#win.geometry('1280x1024')
win.state("zoomed")
style = ttk.Style()

# Configure the style for the scrollbar
style.configure("TScrollbar",
                background="gray",
                troughcolor="light gray",
                gripcount=0,
                gripcolor="white",
                gripinset=2,
                gripborderwidth=0,
                thickness=10)
style.layout("TScrollbar",
             [("Vertical.Scrollbar.trough", {"sticky": "ns"}),
              ("Vertical.Scrollbar.thumb", {"sticky": "ns"})])
data = []
IPpath = ''
arraylist = []
last_plc_value = None
heading_font = tkinter.font.Font(family="Helvetica", size=10, weight="bold")
timems = 10000
readok = False
readvalues = []
file2 = ''
dataheader = ['testheader1', 'testheader2', 'testheader3', 'testheader4', 'testheader5', 'DateTime']
tree = None

#testtags = ["tag1", "tag2", "motor_on", "input1", "output1", "monday", "friday"] # for testing
msg = StringVar()
msg.set('Messages')
custom_font = tkinter.font.Font(size=14)
win.grid_columnconfigure(0, weight=1)
#win.grid_rowconfigure(0, weight=1)

#------------------Frames for easier gui row-column setup------------------------#
frame = Frame(win, height=65, bg='lightgrey', bd=4, relief='raised')
frame.grid(row=0, column=0, pady=10, columnspan=4, sticky='ew')
frame.grid_propagate(False)

frame1 = Frame(win, width=675, height=55, bd=4, relief='raised', bg='lightgrey')
frame1.grid(row=1, column=0, sticky='nw')
frame1.grid_propagate(FALSE)

frame2 = Frame(win, width=700, height=450, bd=3, relief='raised', bg='light blue')#, bg='orange
frame2.grid(row=3, column=0, padx=10, pady=10, sticky='w')
frame2.grid_propagate(False)

frame3 = Frame(win, width=675, height=90, bd=3, relief='raised', bg='light grey')
frame3.grid(row=1, column=0, pady=0, sticky='sw')
frame3.grid_propagate(False)

frame4 = Frame(win,height=200, bg='light green')
frame4.grid(row=5, column=0, padx=10, pady=10, columnspan=3, sticky='ew')
frame4.grid_propagate(False)

frame5 = Frame(win, width=850, height=450, bg='beige', bd=3, relief='raised')
frame5.grid(row=3, column=1, sticky='w')
frame5.grid_propagate(False)

frame6 = Frame(win, height=150, bg='grey', bd=3, relief='raised')
frame6.grid(row=1, column=1, sticky='ew')
frame6.grid_propagate(False)

frame7 = Frame(win,height=50, bg='yellow')
frame7.grid(row=4, column=0, padx=10, pady=10, columnspan=3, sticky='ew')
frame7.grid_propagate(False)

#---------------------------------------------------------------------------------#

def getip():
     global IPpath
     IPpath = Pathip.get()

     if is_valid_ip(IPpath):
        tblbl = Label(frame1, text=f'{IPpath} is Valid and Saved!', font='Arial', fg='Green', bg='lightgrey')
        tblbl.grid(row=0, column=5, sticky='w')
     else:
        tblbl = Label(frame1, text='Invalid IP Address!', font='Arail', fg="Red", bg='lightgrey')
        tblbl.grid(row=0, column=5, sticky='w')

     Pathip.delete(0, END)
 
def is_valid_ip(ip):
    # Regular expression for a valid IPv4 address
    pattern = re.compile(r'^((25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)$')
    return pattern.match(ip) is not None
          
      
def discoverplc():# gets all tags from plc
    global IPpath
    try:
     with LogixDriver(IPpath, init_program_tags=True) as plc:
            # Get all controller and program tags
            tag_list = plc.get_tag_list(program='*')      

           # Insert tags into the text box
            for tag in tag_list:
                text_box_gettags.insert(END, tag['tag_name'] + "\n")
    except Exception as e:
        msg.set(f"Cannot get tags from PLC, check connection. Error: {e}")
     
     
#def testxcel(): for testing change writeexcel on discover button to testxcel
    #msg.set("made it") testing
  #  excelwrite()

#-----------------------------------------------------------------------------------------------------------------#
def read_plc_a(): #Read array from PLC 
  global IPpath, arraylist, timems, readok, last_plc_value

  if not readok: 
        return  # Exit the function without continuing
  intervalms()
  readlinelbl.config(text=" ")
  
  try: 
      array_to_read = entry_Array.get()
      with LogixDriver(IPpath) as plc:
        #array syntax nameofarray{#of elements} example: testarray{5}
        read_plc_array = plc.read(array_to_read) # read 5 elements starting at 0 from an array
        current_plc_value = read_plc_array.value
        if current_plc_value != last_plc_value: 
                arraylist.append(current_plc_value)
                excelwrite()  # Write to Excel only when value changes
                readlinelbl.config(text=", ".join(map(str, arraylist)))  # Update label to show new value
                updated_data = read_excel_data(file2)

               # Display the updated data in the Treeview
                display_data_in_treeview(dataheader, updated_data)
                arraylist.clear()
                last_plc_value = current_plc_value 
        if readok:
          win.after(timems, read_plc_a)
  except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}") 

#-----------------------------------------------------------------------------------------------------------------#
def read_plc_s():#read string from plc
   global IPpath, arraylist, timems, readok, last_plc_value

   if not readok: 
        return  # Exit the function without continuing
      
   intervalms()# get the interval time from TB
   #readlinelbl.config(text=" ")
   
   try:
       string_to_read = entry_string.get()
       with LogixDriver(IPpath) as plc:
         read_plc_string = plc.read(string_to_read)  #read values from plc
         current_plc_value = read_plc_string.value
         if current_plc_value != last_plc_value: 
                arraylist.append(current_plc_value)
                excelwrite()  # Write to Excel only when value changes
               # readlinelbl.config(text=", ".join(map(str, arraylist)))  # Update label to show new value
                updated_data = read_excel_data(file2)

          # Display the updated data in the Treeview
                display_data_in_treeview(dataheader, updated_data)
                arraylist.clear()
                last_plc_value = current_plc_value
               # plc.close() 
         if readok:
           win.after(timems, read_plc_s)
   except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}") 

#-----------------------------------------------------------------------------------------------------------------#
def read_plc_u(): #read udt from plc
    global IPpath, arraylist, timems, readok, last_plc_value, file2

    if not readok:  # Check if logging should stop
        return  # Exit the function without continuing
    
    udt_to_read = entry_udt.get()
    intervalms()
    readlinelbl.config(text=" ") 
    try:      
          with LogixDriver(IPpath) as plc:
            read_udt = plc.read(udt_to_read) 
            if read_udt and hasattr(read_udt, "value"):  # Ensure it's valid
                value_list = list(read_udt.value.values())  # Extract the dictionary and convert to a list
                current_plc_value = value_list
                
                if current_plc_value != last_plc_value:
                  for value in value_list:
                      if isinstance(value, list):  # If value is a list, extend instead of append
                          arraylist.extend(value)    
                      else:
                          arraylist.append(value)
          # Add current datetime to the list
          current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
          arraylist.append(current_datetime)                   
          #print(arraylist) for debugging
          excelwrite()
          # Read the updated Excel data
          updated_data = read_excel_data(file2)

        # Display the updated data in the Treeview
          display_data_in_treeview(dataheader, updated_data)

          readlinelbl.config(text=", ".join(map(str, arraylist))) #print to label on gui to see string being read
          last_plc_value = current_plc_value
          arraylist.clear()
          if readok:
            win.after(timems, read_plc_u)
    except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}")
        
#-----------------------------------------------------------------------------------------------------------------#        
def plc_read_m(): #read multitag from plc
    global IPpath, arraylist, timems, readok, last_plc_value
    if not readok: 
        return  # Exit the function without continuing
    intervalms()
    readlinelbl.config(text="Read Data")
    multilist = []
   
    try:    
         multi_to_read = multi_tag.get()
         multilist = [item.strip() for item in multi_to_read.split(',')]# need to separate list or it thinks its one string!!
       
         with LogixDriver(IPpath) as plc: 
           for i in multilist:     
             multiread = plc.read(i)  # read tags
             current_plc_value = multiread
             
             if current_plc_value != last_plc_value: 
               arraylist.append(multiread.value) # append to our writing list to excel
        #print(arraylist) for debugging
         excelwrite()
         # Read the updated Excel data
         updated_data = read_excel_data(file2)

        # Display the updated data in the Treeview
         display_data_in_treeview(dataheader, updated_data)
         readlinelbl.config(text=", ".join(map(str, arraylist))) #print to label on gui to see string being read
         arraylist.clear()
         last_plc_value = current_plc_value
         if readok:
           win.after(timems, plc_read_m)
    except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}")        
      
#-----------------------------------------------------------------------------------------------------------------#       
def excelwrite():
  global data, file2, arraylist, dataheader
# this is your headers for excel sheet only populates if new sheet started
  dataheader = ['testheader1','testheader2','testheader3','testheader4','testheader5','DateTime']#headers for new sheet

  current_datetime = datetime.now().strftime("%Y_%m_%d")

  #change folder location when get up and running to save location of file to be collecting data.
  #file2 = "C:\\Users\\name\\documents\\test.xlsx" + 
  file2 = "C:\\Users\\PATH GOES HERE\\documents\\test_" + current_datetime + ".xlsx"
  try:
    if os.path.isfile(file2):  # if file already exists append to existing file
     wb = openpyxl.load_workbook(file2)  # load workbook if already exists
     ws = wb.active
     # append the data results to the current excel file
     ws.append(arraylist)
     wb.save(file2)  # save workbook
     wb.close()
    
    else:  # create the excel file if doesn't already exist
        wb = Workbook()
        ws = wb.active
        ws.append(dataheader)
        ws.append(data) 
        wb.save(file2)  # save workbook
        wb.close()
        
  except Exception as e:
        msg.set(f"Cannot connect to Excel. Error: {e}")           
#-----------------------------------------------------------------------------------------------------------------#
#       
def intervalms():
    global timems
    interval = int(timeentryms.get()) 
    if interval > 0:
      timems = interval * 1000
    else:
       timems = 10000 # default 10sec

#-----------------------------------------------------------------------------------------------------------------#      
def stoplogging():
    global readok
    readok = False  
    msg.set(text="STOPPED LOGGING") 
#-----------------------------------------------------------------------------------------------------------------#

def startlogging():
    global readok

    try:
      if entry_udt.get().strip():  # Check if not empty
         readok = True
         read_plc_u()
      if entry_Array.get().strip():  
         readok = True
         read_plc_a()
      if entry_string.get().strip():  
         readok = True
         read_plc_s()
      if multi_tag.get().strip():  
         readok = True
         plc_read_m()
    except Exception as e:
        msg.set(f"Is entry box empty? Error: {e}") 
        
#---------------------------------------------------------------------------------------#

def display_data_in_treeview(headers, data):
  global tree
   # If Treeview already exists, clear the data but don't re-create headers
  if tree:
        for row in tree.get_children():
            tree.delete(row)  # Remove previous rows before inserting new data
  else:
        # Create Treeview only once
        tree = ttk.Treeview(frame4, columns=headers, show='headings', height=5)
        for header in headers:
            tree.heading(header, text=header)
            tree.column(header, width=100)
        tree.pack(expand=True, fill='both')

        # Attach scrollbar only once
        scrollbar = ttk.Scrollbar(frame4, orient="vertical", command=tree.yview, style="TScrollbar")
        scrollbar.pack(side="right", fill="y")
        tree.configure(yscrollcommand=scrollbar.set)

    # Insert new data rows into the existing Treeview
  for row in data:
        tree.insert('', 'end', values=row)
  
#------------------------------------------------------------------------------------------------------#  
def read_excel_data(file_path):
    if not file_path or not os.path.isfile(file_path):  # Check if valid file path
        print(f"Error: File {file_path} does not exist or wasn't created properly.")
        return []

    #print(f"Trying to load: {file_path}")  # Debugging step
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return data

updated_data = read_excel_data(file2)  # Read the written Excel data
display_data_in_treeview(dataheader, updated_data)  # Display updated data in treeview

#-------------------GUI ITEMS------------------------------------------------------------------------#

ipalbl = Label(frame, text="PLC IP Address: ", font='arial' ,bd=2, bg='lightgrey')
ipalbl.grid(row=0, column=0, padx=5, pady = 10, sticky='e')

Pathip = Entry(frame, width=20, font=custom_font)
Pathip.grid(row=0, column=1, padx=15, pady=5, sticky='w')

Saveip = Button(frame,text="Save IP", pady=8, width=15, font='arial', command=getip)
Saveip.grid(row=0, column=2, padx=0, pady=2, sticky='w')


#Tags we are watching for data collection
entry_Array = Entry( frame5, width=60, bg='light blue', font=custom_font)
entry_Array.grid(row=0, column=1, padx=10, pady=30, sticky='nw')

e_arraylabel = Label(frame5, text='syntax: nameofarray{#of elements to read} example: testarray{5}', font= heading_font, bg='beige')
e_arraylabel.grid(row=0, column=1, padx=10, pady=5, sticky='nw')

multi_tag = Entry( frame5, width=60, bg='light blue', font=custom_font )
multi_tag.grid(row=3, column=1, padx=10, pady=25, sticky='w')

multilbl = Label(frame5, text="syntax: tag1, tag2, tag3, etc....", font=heading_font, bg='beige')
multilbl.grid(row=3, column=1, padx=10, pady=0, sticky='nw')

entry_udt = Entry( frame5, width=60, bg='light blue', font=custom_font )
entry_udt.grid(row=2, column=1, padx=10, pady=30, sticky='w')

entry_string = Entry(frame5, width=60, bg='light blue', font=custom_font)
entry_string.grid(row=1, column=1, padx=10, pady=25, sticky='w')

stringlbl = Label(frame5, text="Enter a single tag", font=heading_font, bg='beige')
stringlbl.grid(row=1, column=1, padx=10, sticky='nw')

gettaglbl = Label(frame2, text="Global PLC Tag List", fg='magenta', bg='light blue', font='arial')
gettaglbl.grid(row=0, column=0, padx=150, pady=2, sticky='n')


text_box_gettags = Text(frame2, height=25, width=55, bg='beige')
text_box_gettags.grid(row=0, column=0, columnspan=1, padx=100, pady=26, sticky='w')

# GET ALL plc tags
gettagbtn = Button(frame3, width=15, height=2, text='Get PLC Tags', font='arial', command=discoverplc)
gettagbtn.grid(row=0, column=1, padx=50, pady=15, sticky='ne')

gettaglbl = Label(frame3, text="Get all tags from PLC >>>>>>", font='arial', bg='light grey', fg='blue')
gettaglbl.grid(row=0, column=0, padx=5, pady=25, sticky='w')
# Button to move tag from get to watch
array_button = Button(frame5, width=12, height=2, text="Read Array", font='arial', command=read_plc_a)
array_button.grid(row=0, column=0, padx=5, pady=10, sticky='w')

string_button = Button(frame5, width=12, height=2, text="Read String", font='arial', command=read_plc_s)
string_button.grid(row=1, column=0, padx=5, pady=0, sticky='w')

udt_button = Button(frame5, width=12, height=2, text="UDT", font='arial', command=startlogging)
udt_button.grid(row=2, column=0, padx=5, pady=20, sticky='w')

mult_button = Button(frame5, width=12, height=2, text="Read Multiple", font='arial', command=plc_read_m)
mult_button.grid(row=3, column=0, padx=5, pady=5, sticky='w')

errorlbl = Label(frame, textvariable=msg, width=90, fg='Red', font='arial', bg='lightgrey')
errorlbl.grid(row=0, column=3, padx=20, sticky='w')

stoplogxl = Button(frame5, width=15, height=2, text="Stop Logging", font='arial', command=stoplogging)
stoplogxl.grid(row=5, column=0, padx=5, pady=40, sticky='w')

readlinelbl = Label(frame7, text='Data Read', font='arial')
readlinelbl.grid(row=0, column=0, padx=10, sticky='w')
readlinelbl.config(text=", ".join(map(str, arraylist)))  # Convert list items to a string and print to gui

timeentryms = Entry(frame1, width=10, font=custom_font)
timeentryms.grid(row=0, column=1, padx=2, pady=10, sticky='w')

timelbl = Label(frame1, text="Interval in seconds: ", font='arial', bg='lightgrey')
timelbl.grid(row=0, column=0,padx=5, pady=10, sticky='w')

#------------------------------------------------------------------------------)

win.mainloop()



