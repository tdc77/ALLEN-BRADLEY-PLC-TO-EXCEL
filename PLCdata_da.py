#############################################
#TDC 3-25-25
#
##############################################
from tkinter import *
from tkinter import filedialog, ttk
import tkinter.font
from pycomm3 import LogixDriver
from datetime import datetime
import time
import os
from openpyxl import Workbook
import openpyxl
import tkinter as tk
import pandas as pd
import numpy as np
import ipaddress
import threading
import matplotlib.pyplot as plt


win = Tk()
win.configure(bg='dark grey')
win.title("datacollection")
#win.geometry('1280x1024')
win.state("zoomed")
style = ttk.Style()
thread = None
# Configure the style for the scrollbar
style.configure("Treeview.Scrollbar",
                background="gray",
                troughcolor="light gray",
                gripcount=0,
                gripcolor="white",
                gripinset=2,
                gripborderwidth=0,
                thickness=10)
data = []
IPpath = ''
arraylist = []
last_plc_value = None
heading_font = tkinter.font.Font(family="Helvetica", size=10, weight="bold")
timems = 10000
readok = False
readvalues = []
msg = StringVar()
msg.set('Messages')
cbvar = IntVar()
stop_event = threading.Event()
custom_font = tkinter.font.Font(size=14)
win.grid_columnconfigure(0, weight=1)

#win.grid_rowconfigure(0, weight=1)
# If you have preset column names you can put them here instead of "".  Then you wont have to set columns in program.
columns = {
    "column1": "",
    "column2": "",
    "column3": "",
    "column4": "",
    "column5": "",
    "column6": "",
    "column7": "",
    "column8": ""
}

exfile = StringVar()



#------------------Frames for easier gui row-column setup------------------------#
frame = Frame(win, height=65, bg='lightgrey', bd=4, relief='raised')
frame.grid(row=0, column=0, pady=10, columnspan=4, sticky='ew')
frame.grid_propagate(False)

frame1 = Frame(win, width=675, height=55, bd=4, relief='raised', bg='lightgrey')
frame1.grid(row=1, column=0, sticky='nw')
frame1.grid_propagate(FALSE)

frame2 = Frame(win, width=700, height=450, bd=3, relief='raised', bg='light blue')#, bg='orange
frame2.grid(row=3, column=0, padx=10, pady=10, sticky='w')
frame2.grid_propagate(False)

frame3 = Frame(win, width=675, height=90, bd=3, relief='raised', bg='light grey')
frame3.grid(row=1, column=0, pady=0, sticky='sw')
frame3.grid_propagate(False)

frame4 = Frame(win,height=125, bg='light green')
frame4.grid(row=4, column=0, padx=10, pady=10, columnspan=3, sticky='ew')
frame4.grid_propagate(False)

frame5 = Frame(win, width=850, height=450, bg='beige', bd=3, relief='raised')
frame5.grid(row=3, column=1, sticky='w')
frame5.grid_propagate(False)

frame6 = Frame(win, height=150, bg='grey', bd=3, relief='raised')
frame6.grid(row=1, column=1, sticky='ew')
frame6.grid_propagate(False)

#---------------------------------------------------------------------------------#

def getip():
     global IPpath
     IPpath = Pathip.get()

     if is_valid_ip(IPpath):
        tblbl = Label(frame1, text=f'{IPpath} is Valid and Saved!', font='Arial', fg='Green', bg='lightgrey')
        tblbl.grid(row=0, column=5, sticky='w')
     else:
        tblbl = Label(frame1, text='Invalid IP Address!', font='Arail', fg="Red", bg='lightgrey')
        tblbl.grid(row=0, column=5, sticky='w')

     Pathip.delete(0, END)
 
def is_valid_ip(ip):
     try:
        ipaddress.ip_address(ip)
        return True
     except ValueError:
        return False
   # Regular expression for a valid IPv4 address
   # pattern = re.compile(r'^((25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)\.){3}(25[0-5]|2[0-4][0-9]|[0-1]?[0-9][0-9]?)$')
   # return pattern.match(ip) is not None
          
      
def discoverplc():# gets all tags from plc
    global IPpath
    try:
     with LogixDriver(IPpath, init_program_tags=True) as plc:
            # Get all controller and program tags
            tag_list = plc.get_tag_list(program='*')      

           # Insert tags into the text box
            for tag in tag_list:
                text_box_gettags.insert(END, tag['tag_name'] + "\n")
    except Exception as e:
        msg.set(f"Cannot get tags from PLC, check connection. Error: {e}")
     
     
#def testxcel(): for testing change writeexcel on discover button to testxcel
    #msg.set("made it") testing
  #  excelwrite()

#-----------------------------------------------------------------------------------------------------------------#
def read_plc_a(): #Read array from PLC 
  global IPpath, arraylist, timems, readok, last_plc_value
  
    
  if not readok: 
        msg.set("Waiting to Read")
        return  # Exit the function without continuing
  intervalms()  
  msg.set("Reading array data")
  try: 
      array_to_read = entry_Array.get()
      with LogixDriver(IPpath) as plc:
        #array syntax nameofarray{#of elements} example: testarray{5}
        read_plc_array = plc.read(array_to_read) # read 5 elements starting at 0 from an array
        current_plc_value = read_plc_array.value
        if current_plc_value != last_plc_value: # Write to Excel only when value changes
          if isinstance(current_plc_value, list):  # Ensure correct structure
            arraylist.extend(current_plc_value)  # Extend list instead of appending it as a nested item
          else:
            arraylist.append(current_plc_value)
        excelwrite()  
        arraylist.clear()
        last_plc_value = current_plc_value 
        if readok:
          win.after(timems, read_plc_a)
  except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}") 

#-----------------------------------------------------------------------------------------------------------------#
def read_plc_s():#read string from plc
   global IPpath, arraylist, timems, readok, last_plc_value

   
   if not readok: 
        return  # Exit the function without continuing
      
   intervalms()# get the interval time from TB
   msg.set("Reading single string")
   try:
       string_to_read = entry_string.get()
       with LogixDriver(IPpath) as plc:
         read_plc_string = plc.read(string_to_read)  #read values from plc
         current_plc_value = read_plc_string.value
         if current_plc_value != last_plc_value:  # Write to Excel only when value changes
                arraylist.append(current_plc_value)
                excelwrite()         
                arraylist.clear()
                last_plc_value = current_plc_value
               # plc.close() 
         if readok:
           win.after(timems, read_plc_s)
   except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}") 

#-----------------------------------------------------------------------------------------------------------------#
def read_plc_u(): #read udt from plc
    global IPpath, arraylist, timems, readok, last_plc_value

   
    if not readok:  # Check if logging should stop
        return  # Exit the function without continuing
    
    udt_to_read = entry_udt.get()
    intervalms()
    msg.set("Reading UDT Data")
    #readlinelbl.config(text=" ") 
    try:     
          with LogixDriver(IPpath) as plc:
            read_udt = plc.read(udt_to_read) 
            if read_udt and hasattr(read_udt, "value"):  # Ensure it's valid
                value_list = list(read_udt.value.values())  # Extract the dictionary and convert to a list
                current_plc_value = value_list
                
                if current_plc_value != last_plc_value:
                  for value in value_list:
                      if isinstance(value, list):  # If value is a list, extend instead of append
                          arraylist.extend(value)    
                      else:
                          arraylist.append(value)
          # Add current datetime to the list
          current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
          arraylist.append(current_datetime)                   
          #print(arraylist) for debugging
          excelwrite()
         # readlinelbl.config(text=", ".join(map(str, arraylist))) #print to label on gui to see string being read
          last_plc_value = current_plc_value
          arraylist.clear()
          if readok:
            win.after(timems, read_plc_u)
    except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}")
        
#-----------------------------------------------------------------------------------------------------------------#        
def plc_read_m(): #read multitag from plc
    global IPpath, arraylist, timems, readok, last_plc_value

    
    if not readok: 
        return  # Exit the function without continuing
    intervalms()
    #readlinelbl.config(text=" ")
    msg.set("Reading multiple tag data")
    multilist = []
   
    try:    
         multi_to_read = multi_tag.get()
         multilist = [item.strip() for item in multi_to_read.split(',')]# need to separate list or it thinks its one string!!
       
         with LogixDriver(IPpath) as plc: 
           for i in multilist:     
             multiread = plc.read(i)  # read tags
             current_plc_value = multiread
             
             if current_plc_value != last_plc_value: 
               arraylist.append(multiread.value) # append to our writing list to excel
        #print(arraylist) for debugging
         excelwrite()
         arraylist.clear()
         last_plc_value = current_plc_value
         if readok:
           win.after(timems, plc_read_m)
    except Exception as e:
        msg.set(f"Cannot connect to PLC. Error: {e}")        
      
#-----------------------------------------------------------------------------------------------------------------#       
def excelwrite():
  global data, file2, arraylist, dataheader
# this is your headers for excel sheet only populates if new sheet started
  dataheader = list(columns.values())
  current_datetime = datetime.now().strftime("%Y_%m_%d")

  #change folder location when get up and running to save location of file to be sent
   
  file2 = "C:\\Users\\terry\\documents\\test_" + current_datetime + ".xlsx" # change this to your path
  try:
    if os.path.isfile(file2):  # if file already exists append to existing file
     wb = openpyxl.load_workbook(file2)  # load workbook if already exists
     ws = wb.active
     # append the data results to the current excel file
     ws.append(arraylist)
     wb.save(file2)  # save workbook
     wb.close()
     #print("Headers:", dataheader) Testing only
     #print("Data:", data) Testing only
     display_data_in_treeview(dataheader,[tuple(arraylist)], frame4)
    
    else:  # create the excel file if doesn't already exist
        wb = Workbook()
        ws = wb.active
        ws.append(dataheader)
        ws.append(data) 
        wb.save(file2)  # save workbook
        wb.close()
        display_data_in_treeview(dataheader,[tuple(arraylist)], frame4)
  except Exception as e:
        msg.set(f"Cannot connect to Excel. Error: {e}")           
#-----------------------------------------------------------------------------------------------------------------#
#       
def intervalms():
    global timems
    interval = int(timeentryms.get()) 
    if interval > 0:
      timems = interval * 1000
    else:
       timems = 10000 # default 10sec

#-----------------------------------------------------------------------------------------------------------------#      
def stoplogging():
    global readok
    readok = False  
    msg.set("STOPPED LOGGING") 
#-----------------------------------------------------------------------------------------------------------------#

def startlogging():
    global readok

    try:
      if entry_udt.get().strip():  # Check if not empty
         readok = True
         read_plc_u()
      if entry_Array.get().strip():  
         readok = True
         read_plc_a()
      if entry_string.get().strip():  
         readok = True
         read_plc_s()
      if multi_tag.get().strip():  
         readok = True
         plc_read_m()
    except Exception as e:
        msg.set(f"Is entry box empty? Error: {e}") 
        
#------------------------------------------------------------------------------------------------#

def checktime():
    global thread
    def run():
      msg.set("Midnight checker started.")
      win.after(5000, lambda: msg.set(""))
      while not stop_event.is_set():
        now = datetime.now()
        if now.hour == 0 and now.minute == 0:
           msg.set("Creating new sheet...")
           createnewsheet()
           win.after(3000, lambda: msg.set(""))
           time.sleep(60) # Wait a minute to avoid multiple triggers
        time.sleep(1) # Check every second

    if cbvar.get() == 1: 
      if thread is None or not thread.is_alive():
          stop_event.clear() # Reset the stop flag
          thread = threading.Thread(target=run, daemon=True)
          thread.start()
    elif cbvar.get() == 0:
         stop_event.set()
         msg.set("No new sheet will be made")
         win.after(5000, lambda: msg.set(""))


 #---------------------------------------------------------------------------------------#   
    
def createnewsheet():
    global  file2, dataheader
    # this is your headers for excel sheet only populates if new sheet started
    dataheader = list(columns.values())
    current_datetime = datetime.now().strftime("%Y_%m_%d")

    #change folder location when get up and running to save location of file to be sent
    file2 = "C:\\Users\\name\\documents\\test_" + current_datetime + ".xlsx"    
    if not os.path.isfile(file2):
        wb = Workbook()
        ws = wb.active
        ws.append(dataheader)
        wb.save(file2)  # save workbook
        wb.close()
        
#-------------------------------------------------------------------------------------#

def display_data_in_treeview(headers, data, frame4):
    if not hasattr(frame4, "treeview"):  # Initialize Treeview only once
        frame4.treeview = ttk.Treeview(frame4, columns=headers, show="headings", height=5)
        frame4.treeview.pack(side="left", expand=True, fill="both")

        for header in headers:
            frame4.treeview.heading(header, text=header)
            frame4.treeview.column(header, width=100)

        scrollbar = ttk.Scrollbar(frame4, orient="vertical", command=frame4.treeview.yview)
        frame4.treeview.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

    # **DO NOT clear rows here** — Instead, just append new data
    for row in data:
        frame4.treeview.insert("", "end", values=row)
#-----------------------------------------------------------------------------------------------------------------#  
def update_combobox_options():
    headerlist['values'] = list(columns.keys())

#-------------------------------------------------------------------------------------------------------#
def change_header_name():
      try:
        selected_column = headerlist.get()
        new_header_name = headername.get() 
        if selected_column in columns:
          columns[selected_column] = new_header_name
        #print(columns) testing only
        headername.delete(0, END)
      except Exception as e:
        msg.set("Error in Naming columns Error:{e}")  
  
#--------------------------------------------------------------------------------------------------------#        
def calculate_cpk(tree, label_result, uslentry, lslentry):
    """ Calculate CpK from the column data stored in the Treeview """
    values = []
    
    try:
        
        usl = float(uslentry.get())
        lsl = float(lslentry.get())

        # Extract column data from Treeview
        for row in tree.get_children():
            row_data = tree.item(row)["values"]  # Get row values
            if row_data:
                values.append(float(row_data[1]))  # Convert values to float

        if values:
            mean_value = np.mean(values)
            std_dev = np.std(values, ddof=1)  # Sample standard deviation

            # Calculate CpK
            cpk = min((usl - mean_value) / (3 * std_dev), (mean_value - lsl) / (3 * std_dev))

            label_result.config(text=f"CpK: {cpk:.3f}")  # Display result
        else:
            label_result.config(text="No valid data found.")
    
    except ValueError:
        label_result.config(text="Invalid USL or LSL! Please enter numeric values.")
    
    except Exception as e:
        label_result.config(text=f"Error: {e}")

#------------------------------------------------------------------------------------------------------------#                
def calculate_sum(tree, label_result):
    """ Calculate the sum of the column data stored in the Treeview """
    values = []

  
    for row in tree.get_children():
        row_data = tree.item(row)["values"]  # Get row values
        if row_data:  # Ensure there is data
            values.append(float(row_data[1]))  # Convert values to float for sum calculation

    if values:
        column_sum = sum(values)  # Compute sum
        label_result.config(text=f"Sum: {column_sum:.2f}")  # Update label with result
    else:
        label_result.config(text="No valid data found.")
        
#-------------------------------------------------------------------------------------------------------------#
def calculate_avg(tree, label_result):
   
    values = []

    # Extract column data from Treeview
    for row in tree.get_children():
        row_data = tree.item(row)["values"]  # Get row values
        if row_data:  # Ensure there is data
            values.append(float(row_data[1]))  # Convert values to float for math operations

    if values:
        column_avg = sum(values) / len(values)  # Compute average
        label_result.config(text=f"Average: {column_avg:.2f}")  # Update label with result
    else:
        label_result.config(text="No valid data found.")

 #-------------------------------------------------------------------------------------------------------------#       
        
def load_column_data(exfile, column_name, tree, label_result):
   
    exfilepath = exfile.get()  # Retrieve file path from StringVar

    if exfilepath:
        df = pd.read_excel(exfilepath)  # Load Excel file

        # Clear previous data in Treeview
        for row in tree.get_children():
            tree.delete(row)

        if column_name in df.columns:
            column_data = df[column_name]  # Extract column data

            # Insert data into Treeview
            for i, value in enumerate(column_data):
                tree.insert("", "end", values=(i+1, value))  # Adding row index and column value
        else:
            label_result.config("Invalid column name. Please choose from the headers above.")
    else:
         label_result.config("No file selected.")


#-----------------------------------------------------------------------------------------#
import matplotlib.pyplot as plt

def plot_graph(tree, num_points_entry):
    values = []

    # Extract column data from Treeview
    for row in tree.get_children():
        row_data = tree.item(row)["values"]
        if row_data:
            try:
                values.append(float(row_data[1]))
            except ValueError:
                continue  # Skip rows with non-numeric data

    # Get number of points to plot from Entry widget
    try:
        num_points = int(num_points_entry.get())
        if num_points < 1:
            raise ValueError
    except ValueError:
        print("Please enter a valid positive integer for number of points.")
        return

    # Limit the number of points to plot
    values_to_plot = values[:num_points]

    if values_to_plot:
        plt.figure(figsize=(6, 4))
        plt.plot(range(len(values_to_plot)), values_to_plot, marker='o', linestyle='-', color='blue', label="Values")
        plt.xlabel("Index")
        plt.ylabel("Value")
        plt.title(f"Data Plot (First {num_points} Points)")
        plt.legend()
        plt.show()
    else:
        print("No valid data found.")

        
#----------------------------------------------------------------------------------------------------------------------------#    
def datascreen():
    """ Creates the new window with UI elements """
    new_window = tk.Toplevel(win)
    new_window.title("New GUI Window")
    new_window.geometry('800x800')
    # Keep focus on this window so it doesnt minimize automatically
    new_window.grab_set()

    
    exfile = tk.StringVar()

    def open_file():
        
        exfilepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
        
        if exfilepath:
            df = pd.read_excel(exfilepath)  
            headers = list(df.columns)  # Get headers
            label.config(text="Headers:\n" + "\n".join(headers))  # Display headers
            
            exfile.set(exfilepath)  # Set file path in StringVar
            
#--------------------------NEW GUI ITEMS -------------------------------------------#
    frame_result = Frame(new_window, width=400, height=50)  # Fixed size
    frame_result.grid(row=8, column=0, padx=10, pady=10, sticky='w')
    
    frame_header = Frame(new_window, width=250, height=50)
    frame_header.grid(row=0, column=0, pady=10, sticky='w')
    
    dataframe = Frame(new_window, width=400, height=100)
    dataframe.grid(row=1, column=0, sticky='w')
    
    limitframe = Frame(new_window, width=600, height=80)
    limitframe.grid(row=6, column=0, sticky='w')

    label_result = Label(frame_result, text='', bg='yellow', font='arial')
    label_result.pack(fill="both", expand=False)  # Prevent resizing issues

    # Labels and Entry widgets
    label = Label(frame_header, text="Select an Excel file to view headers:")
    label.grid(row=0, column=0, padx=10, pady=10)

    entry_file = Entry(dataframe, width=40, textvariable=exfile)  # Entry using StringVar
    entry_file.grid(row=0, column=1, pady=10, sticky='w')

    open_button = Button(frame_header, text="Open Excel File", command=open_file)
    open_button.grid(row=0, column=1, padx=10, pady=10)

    datacolumn = Entry(dataframe, width=45)
    datacolumn.grid(row=1, column=1, pady=10, sticky='w')

    lblcolumn = Label(dataframe, text='Column to analyze: ')
    lblcolumn.grid(row=1, column=0)
    
    btngetdata = Button(dataframe, text="Get Column", command=lambda: load_column_data(exfile, datacolumn.get(), tree, label_result))
    btngetdata.grid(row=1, column=2, padx=10, pady=10)

    exlbl = Label(dataframe, text="Excel File:")
    exlbl.grid(row=0, column=0, padx=5, pady=10, sticky='w')

    close_button = Button(new_window, width=15, text="Close", command=new_window.destroy)
    close_button.grid(row=5, column=0, padx=10, pady=10)
    
    uslentry = Entry(limitframe, width=10)
    uslentry.grid(row=1, column=0, padx=55, pady=10, sticky='w')
    
    lslentry = Entry(limitframe, width=10)
    lslentry.grid(row=1, column=1, padx=40, pady=10, sticky='e')
    
    lblusl = Label(limitframe, text='CPK USL')
    lblusl.grid(row=0, column=0, padx=10, pady=1,sticky='s')
    
    lbllsl = Label(limitframe, text='CPK LSL')
    lbllsl.grid(row=0, column=1, padx=10, pady=1, sticky='s')
    
    # Create Treeview Table
    tree = ttk.Treeview(new_window, columns=("Index", "Column Data"), show="headings")
    tree.heading("Index", text="Index")
    tree.heading("Column Data", text="Column Data")
    tree.grid(row=7, column=0, columnspan=2)
    
    avgbtn = Button(new_window, width=15, text="Average", command=lambda: calculate_avg(tree, label_result))
    avgbtn.grid(row=7, column=3, sticky='n')
    
    sumbtn = Button(new_window, width=15, text='Sum', command=lambda: calculate_sum(tree, label_result))
    sumbtn.grid(row=7, column=3, sticky='w')
    
    cpkbtn = Button(new_window, width=15, text='CPK', command=lambda: calculate_cpk(tree, label_result, uslentry, lslentry))
    cpkbtn.grid(row=7, column=3,  sticky='s')

    graphbtn = Button(new_window, width=15, text="Graph", command=lambda: plot_graph(tree, num_points_entry))
    graphbtn.grid(row=7, column=3, pady=45, sticky='s')

    num_points_entry = Entry(limitframe, width=5)
    num_points_entry.grid(row=1, column=4, sticky='w')
    
    nplbl = Label(limitframe, text="number of points to graph:")
    nplbl.grid(row=1, column=3, padx=10, sticky='w')
#-------------------ROOT GUI ITEMS------------------------------------------------------------------------#

ipalbl = Label(frame, text="PLC IP Address: ", font='arial' ,bd=2, bg='lightgrey')
ipalbl.grid(row=0, column=0, padx=5, pady = 10, sticky='e')

Pathip = Entry(frame, width=20, font=custom_font)
Pathip.grid(row=0, column=1, padx=15, pady=5, sticky='w')

Saveip = Button(frame,text="Save IP", pady=8, width=15, font='arial', command=getip)
Saveip.grid(row=0, column=2, padx=0, pady=2, sticky='w')

#Tags we are watching for data collection
entry_Array = Entry( frame5, width=60, bg='light blue', font=custom_font)
entry_Array.grid(row=0, column=1, padx=10, pady=30, sticky='nw')

e_arraylabel = Label(frame5, text='syntax: nameofarray{#of elements to read} example: testarray{5}', font= heading_font, bg='beige')
e_arraylabel.grid(row=0, column=1, padx=10, pady=5, sticky='nw')

multi_tag = Entry( frame5, width=60, bg='light blue', font=custom_font )
multi_tag.grid(row=3, column=1, padx=10, pady=25, sticky='w')

multilbl = Label(frame5, text="syntax: tag1, tag2, tag3, etc....", font=heading_font, bg='beige')
multilbl.grid(row=3, column=1, padx=10, pady=0, sticky='nw')

entry_udt = Entry( frame5, width=60, bg='light blue', font=custom_font )
entry_udt.grid(row=2, column=1, padx=10, pady=30, sticky='w')

entry_string = Entry(frame5, width=60, bg='light blue', font=custom_font)
entry_string.grid(row=1, column=1, padx=10, pady=25, sticky='w')

stringlbl = Label(frame5, text="Enter a single tag", font=heading_font, bg='beige')
stringlbl.grid(row=1, column=1, padx=10, sticky='nw')

gettaglbl = Label(frame2, text="Global PLC Tag List", fg='magenta', bg='light blue', font='arial')
gettaglbl.grid(row=0, column=0, padx=150, pady=2, sticky='n')

text_box_gettags = Text(frame2, height=25, width=55, bg='beige')
text_box_gettags.grid(row=0, column=0, columnspan=1, padx=100, pady=26, sticky='w')

# GET ALL plc tags
gettagbtn = Button(frame3, width=15, height=2, text='Get PLC Tags', font='arial', command=discoverplc)
gettagbtn.grid(row=0, column=1, padx=50, pady=15, sticky='ne')

gettaglbl = Label(frame3, text="Get all tags from PLC >>>>>>", font='arial', bg='light grey', fg='blue')
gettaglbl.grid(row=0, column=0, padx=5, pady=25, sticky='w')
# Button to move tag from get to watch
array_button = Button(frame5, width=12, height=2, text="Read Array", font='arial', command=startlogging)
array_button.grid(row=0, column=0, padx=5, pady=10, sticky='w')

string_button = Button(frame5, width=12, height=2, text="Read String", font='arial', command=startlogging)
string_button.grid(row=1, column=0, padx=5, pady=0, sticky='w')

udt_button = Button(frame5, width=12, height=2, text="UDT", font='arial', command=startlogging)
udt_button.grid(row=2, column=0, padx=5, pady=20, sticky='w')

mult_button = Button(frame5, width=12, height=2, text="Read Multiple", font='arial', command=startlogging)
mult_button.grid(row=3, column=0, padx=5, pady=5, sticky='w')

errorlbl = Label(frame, textvariable=msg, width=90, fg='Red', font='arial', bg='lightgrey')
errorlbl.grid(row=0, column=3, padx=20, sticky='w')

stoplogxl = Button(frame5, width=15, height=2, text="Stop Logging", font='arial', command=stoplogging)
stoplogxl.grid(row=5, column=0, padx=5, pady=40, sticky='w')

databtn = Button(frame6, width=15, height=2, text="Data", command=datascreen)
databtn.grid(row=2, column=3, padx=15, sticky='w')

timeentryms = Entry(frame1, width=10, font=custom_font)
timeentryms.grid(row=0, column=1, padx=2, pady=10, sticky='w')

timelbl = Label(frame1, text="Interval in seconds: ", font='arial', bg='lightgrey')
timelbl.grid(row=0, column=0,padx=5, pady=10, sticky='w')

headerlist = ttk.Combobox(frame6, width=25, postcommand=update_combobox_options)
headerlist['state'] = 'readonly'
headerlist.grid(row=2, column=0, padx=3, pady=5, sticky='n')

headername = Entry(frame6, width=35)
headername.grid(row=2, column=1, padx=5, pady=5, sticky='n')

namechgbtn = Button(frame6, text="Change", width=15, height=2, command= change_header_name)
namechgbtn.grid(row=2, column=2, padx=10, pady=0, sticky='n')

lblcombobox = Label(frame6, text="Header Name Change")
lblcombobox.grid(row=1, column=0, padx=25, pady=10, sticky='w')

lblnewname = Label(frame6, text='New Name')
lblnewname.grid(row=1, column=1, padx=150, pady=10, sticky='w')

lbltitlef6 = Label(frame6, text="Add names to columns for excel sheet", font='arial')
lbltitlef6.grid(row=0, column=1, padx=25, pady=9, sticky='n')

cbnewsheet = Checkbutton(frame5, text="Create new sheet at midnight", variable=cbvar, command=checktime)
cbnewsheet.grid(row=5, column=1, padx=50, sticky='w')
#------------------------------------------------------------------------------)


     
win.mainloop()



